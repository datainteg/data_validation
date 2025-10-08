# validator_core.py
# Author: Akshay + Copilot
# Purpose: Compare & validate CSV/Excel (PROD vs UAT) and export styled Excel.

import io
import re
import math
from pathlib import Path
from typing import List, Optional, Dict, Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


class CompareConfig:
    def __init__(
        self,
        # Header / parsing
        case_insensitive_headers: bool = True,
        strip_whitespace: bool = True,

        # Rendering
        placeholder: str = "N/A",
        thousand_sep: str = ",",
        decimal_sep: str = ".",

        # Comparison tolerances
        numeric_abs_tol: float = 0.0,
        numeric_rel_tol: float = 0.0,
        date_tolerance_days: int = 0,

        # Missing handling & output mode
        treat_empty_as_missing: bool = True,
        output_mode: str = "all_rows",     # "all_rows" | "mismatches_only"

        # Key detection
        max_key_combo: int = 2,
        key_uniqueness_threshold: float = 0.90,

        # Type detection
        detect_dates: bool = True,
        detect_numbers: bool = True,

        # Preview limit (UI)
        limit_preview_rows: int = 500,

        # NEW: missing equivalence & one-sided rows
        missing_tokens: Optional[List[str]] = None,   # tokens treated as missing (case-insensitive)
        equalize_missing_tokens: bool = True,         # blank/N/A/NA/NaN are equal
        suppress_missing_side_row: bool = True,       # << default TRUE: do not print N/A line for missing side
        ignore_missing_records: bool = False,         # drop one-sided pairs entirely
        left_label: str = "PROD",                     # Source label (left file)
        right_label: str = "UAT",                     # Source label (right file)
    ):
        self.case_insensitive_headers = case_insensitive_headers
        self.strip_whitespace = strip_whitespace
        self.placeholder = placeholder
        self.thousand_sep = thousand_sep
        self.decimal_sep = decimal_sep
        self.numeric_abs_tol = numeric_abs_tol
        self.numeric_rel_tol = numeric_rel_tol
        self.date_tolerance_days = date_tolerance_days
        self.treat_empty_as_missing = treat_empty_as_missing
        self.output_mode = output_mode
        self.max_key_combo = max_key_combo
        self.key_uniqueness_threshold = key_uniqueness_threshold
        self.detect_dates = detect_dates
        self.detect_numbers = detect_numbers
        self.limit_preview_rows = limit_preview_rows

        self.missing_tokens = missing_tokens if missing_tokens is not None else ["", "N/A", "NA"]
        self.equalize_missing_tokens = equalize_missing_tokens
        self.suppress_missing_side_row = suppress_missing_side_row
        self.ignore_missing_records = ignore_missing_records
        self.left_label = left_label
        self.right_label = right_label


# ---------- Utilities ----------
ID_LIKE_RE = re.compile(r"\b(id|code|key|number|no|acct|account)\b", re.I)
NUM_PATTERN = re.compile(r"^[+\-]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?$")


def normalize_headers(df: pd.DataFrame, cfg: CompareConfig) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame):
        raise ValueError(f"normalize_headers expected DataFrame, got {type(df)}")
    df = df.copy()
    cols = df.columns.astype(str).str.strip()
    if cfg.case_insensitive_headers:
        cols = cols.str.lower()
    df.columns = cols
    return df


def _read_excel(obj, sheet_name):
    try:
        return pd.read_excel(obj, sheet_name=sheet_name or 0, dtype=str, engine="openpyxl")
    except TypeError:
        return pd.read_excel(obj, sheet_name=sheet_name or 0, dtype=str)


def load_any(source: Any, cfg: CompareConfig, sheet: Optional[str] = None) -> pd.DataFrame:
    # Reset pointer for file-like
    if hasattr(source, "seek"):
        try:
            source.seek(0)
        except Exception:
            pass

    if isinstance(source, (str, Path)):
        p = str(source).lower()
        if p.endswith((".csv", ".txt")):
            df = pd.read_csv(source, dtype=str, keep_default_na=False)
        else:
            df = _read_excel(source, sheet)
    else:
        name = getattr(source, "name", "uploaded").lower()
        if name.endswith((".csv", ".txt")):
            df = pd.read_csv(source, dtype=str, keep_default_na=False)
        else:
            df = _read_excel(source, sheet)

    if df is None:
        raise ValueError("load_any: read returned None (not a DataFrame).")
    if isinstance(df, dict):  # dict of sheets → pick first
        df = df[next(iter(df))]

    df = normalize_headers(df, cfg)
    if cfg.strip_whitespace:
        df = df.apply(lambda s: s.str.strip() if s.dtype == object else s)
    if cfg.treat_empty_as_missing:
        df = df.fillna("")
    return df


def is_missing(val: Any, cfg: CompareConfig) -> bool:
    """Treat np.nan/None/'' and configured tokens as missing."""
    # Catch real NaN/None early
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return True
    try:
        if pd.isna(val):
            return True
    except Exception:
        pass
    s = str(val).strip()
    if s == "":
        return True
    tokens = {t.lower() for t in (cfg.missing_tokens or []) if t is not None}
    return s.lower() in tokens


def looks_like_number_string(s: str) -> bool:
    return bool(NUM_PATTERN.fullmatch(str(s)))


def has_leading_zero(s: str) -> bool:
    s = str(s)
    return (len(s) > 1) and s.startswith("0") and s[1:].isdigit()


def to_float_safe(s: str, thousand_sep: str = ",", decimal_sep: str = "."):
    try:
        s2 = str(s)
        if thousand_sep:
            s2 = s2.replace(thousand_sep, "")
        if decimal_sep != ".":
            s2 = s2.replace(decimal_sep, ".")
        return float(s2)
    except Exception:
        return None


def detect_numeric_cols(df1: pd.DataFrame, df2: pd.DataFrame, exclude: List[str], cfg: CompareConfig) -> List[str]:
    if not cfg.detect_numbers:
        return []
    numeric = []
    for col in (set(df1.columns) | set(df2.columns)) - set(exclude):
        vals = pd.concat([df1.get(col, pd.Series(dtype=str)), df2.get(col, pd.Series(dtype=str))], ignore_index=True)
        vals = vals[vals.notna() & (vals != "")]
        if vals.empty:
            continue
        if any(has_leading_zero(v) for v in vals.head(1000).astype(str)):
            continue
        sample = vals.head(1000).astype(str)
        if all(looks_like_number_string(v) for v in sample):
            numeric.append(col)
    return sorted(numeric)


def detect_date_cols(df1: pd.DataFrame, df2: pd.DataFrame, exclude: List[str], cfg: CompareConfig) -> List[str]:
    if not cfg.detect_dates:
        return []
    date_cols = []
    probe = (set(df1.columns) | set(df2.columns)) - set(exclude)
    for col in probe:
        vals = pd.concat([df1.get(col, pd.Series(dtype=str)), df2.get(col, pd.Series(dtype=str))], ignore_index=True)
        vals = vals[vals.notna() & (vals != "")]
        if vals.empty:
            continue
        parsed = pd.to_datetime(vals.head(300).astype(str), errors="coerce")
        if parsed.notna().mean() >= 0.80:
            date_cols.append(col)
    return sorted(date_cols)


def pick_keys_auto(df1: pd.DataFrame, df2: pd.DataFrame, cfg: CompareConfig) -> Optional[List[str]]:
    common = [c for c in df1.columns if c in df2.columns]
    if not common:
        return None

    cand = []
    for c in common:
        s1 = df1[c].replace("", pd.NA).dropna()
        s2 = df2[c].replace("", pd.NA).dropna()
        if len(s1) == 0 or len(s2) == 0:
            continue
        r1 = s1.nunique() / max(1, len(s1))
        r2 = s2.nunique() / max(1, len(s2))
        if r1 >= cfg.key_uniqueness_threshold and r2 >= cfg.key_uniqueness_threshold:
            cand.append((c, (r1 + r2) / 2.0))

    id_like = [c for c in common if ID_LIKE_RE.search(c)]

    if cand:
        # prefer ID-like among candidates
        id_like_cands = [t for t in cand if t[0] in id_like]
        (id_like_cands or cand).sort(key=lambda x: (-x[1], x[0]))
        return [(id_like_cands or cand)[0][0]]

    if id_like:
        return [id_like[0]]

    if cfg.max_key_combo >= 2:
        best, best_score = None, -1.0
        for i, c1 in enumerate(common):
            for c2 in common[i + 1:]:
                s1 = (df1[c1] + "||" + df1[c2]).replace("", pd.NA).dropna()
                s2 = (df2[c1] + "||" + df2[c2]).replace("", pd.NA).dropna()
                if len(s1) == 0 or len(s2) == 0:
                    continue
                r1 = s1.nunique() / max(1, len(s1))
                r2 = s2.nunique() / max(1, len(s2))
                score = (r1 + r2) / 2.0
                if score > best_score:
                    best_score = score
                    best = [c1, c2]
        if best and best_score >= (cfg.key_uniqueness_threshold - 0.05):
            return best

    return None


def add_key_and_seq(df: pd.DataFrame, keys: Optional[List[str]]):
    if keys is None:
        df = df.copy()
        df["__row_order__"] = range(1, len(df) + 1)
        df["__key__"] = df["__row_order__"].astype(str)
        df["__seq__"] = 0
        return df, ["__row_order__"]
    df = df.copy()
    df["__key__"] = df[keys].astype(str).agg("|".join, axis=1)
    df["__seq__"] = df.groupby("__key__").cumcount()
    return df, keys


def build_pairs(df1k: pd.DataFrame, df2k: pd.DataFrame) -> pd.DataFrame:
    cols1 = ["__key__", "__seq__"] + [c for c in df1k.columns if c not in ["__key__", "__seq__"]]
    cols2 = ["__key__", "__seq__"] + [c for c in df2k.columns if c not in ["__key__", "__seq__"]]
    merged = pd.merge(df1k[cols1], df2k[cols2], on=["__key__", "__seq__"], how="outer", suffixes=("_f1", "_f2"))
    return merged.sort_values(["__key__", "__seq__"]).reset_index(drop=True)


def numeric_equal(a, b, cfg: CompareConfig) -> bool:
    if cfg.equalize_missing_tokens and is_missing(a, cfg) and is_missing(b, cfg):
        return True
    fa = to_float_safe(a, cfg.thousand_sep, cfg.decimal_sep) if not is_missing(a, cfg) else None
    fb = to_float_safe(b, cfg.thousand_sep, cfg.decimal_sep) if not is_missing(b, cfg) else None
    if fa is None or fb is None:
        return (str(a).strip() == str(b).strip())
    if cfg.numeric_abs_tol == 0.0 and cfg.numeric_rel_tol == 0.0:
        return fa == fb
    return math.isclose(fa, fb, rel_tol=cfg.numeric_rel_tol, abs_tol=cfg.numeric_abs_tol)


def date_equal(a, b, cfg: CompareConfig) -> bool:
    if cfg.equalize_missing_tokens and is_missing(a, cfg) and is_missing(b, cfg):
        return True
    da = pd.to_datetime(a, errors="coerce")
    db = pd.to_datetime(b, errors="coerce")
    if pd.isna(da) or pd.isna(db):
        return (str(a).strip() == str(b).strip())
    if cfg.date_tolerance_days <= 0:
        return da == db
    return abs((da.normalize() - db.normalize()).days) <= cfg.date_tolerance_days


def cell_equal(a, b, col, numeric_cols, date_cols, cfg: CompareConfig) -> bool:
    if cfg.equalize_missing_tokens and is_missing(a, cfg) and is_missing(b, cfg):
        return True
    if col in numeric_cols:
        return numeric_equal(a, b, cfg)
    if col in date_cols:
        return date_equal(a, b, cfg)
    return (str(a).strip() == str(b).strip())


def write_excel(
    merged: pd.DataFrame,
    display_cols: List[str],
    numeric_cols: List[str],
    date_cols: List[str],
    keys_used: Optional[List[str]],
    cfg: CompareConfig,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Results"

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_left = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")   # PROD
    fill_right = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # UAT
    fill_mismatch = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    fill_yes = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_no = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    header = ["Source"] + display_cols + ["RecordStatus", "Validation"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def write_cell(cell, value, col_name):
        # Missing displayed as blank
        if is_missing(value, cfg):
            cell.value = ""
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="left")
            return
        # Detect numerics for right aligned display (preserve IDs with leading zeros)
        if col_name in numeric_cols and looks_like_number_string(value) and not has_leading_zero(value):
            num = to_float_safe(value, cfg.thousand_sep, cfg.decimal_sep)
            if num is not None and not math.isnan(num):
                cell.value = num
                cell.number_format = "#,##0.########"
                cell.alignment = Alignment(horizontal="right")
                return
        cell.value = str(value)
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left")

    # Iterate pairs → optionally suppress missing side
    for _, r in merged.iterrows():
        left_vals = {c: r.get(f"{c}_f1", "") for c in display_cols}
        right_vals = {c: r.get(f"{c}_f2", "") for c in display_cols}

        # Missing row flags (all display cols missing on a side)
        left_missing_row = all(is_missing(r.get(f"{c}_f1", ""), cfg) for c in display_cols)
        right_missing_row = all(is_missing(r.get(f"{c}_f2", ""), cfg) for c in display_cols)

        # Ignore one-sided pairs entirely if asked
        if cfg.ignore_missing_records and (left_missing_row ^ right_missing_row):
            continue

        # Compute equality when both exist
        both_present = (not left_missing_row) and (not right_missing_row)
        all_equal = True
        if both_present:
            for col in display_cols:
                a = r.get(f"{col}_f1", "")
                b = r.get(f"{col}_f2", "")
                if not cell_equal(a, b, col, numeric_cols, date_cols, cfg):
                    all_equal = False
                    break

        # Mismatches-only skip for full matches (keep one-sided as they are interesting)
        if cfg.output_mode == "mismatches_only" and both_present and all_equal:
            continue

        # Decide which rows to emit
        emit_left = not left_missing_row
        emit_right = not right_missing_row

        if cfg.suppress_missing_side_row:
            # Only show the present side for one-sided pairs
            if left_missing_row and not right_missing_row:
                emit_left = False; emit_right = True
            elif right_missing_row and not left_missing_row:
                emit_left = True; emit_right = False

        # RecordStatus value
        rec_status = (
            "BOTH" if both_present
            else ("ONLY IN PROD" if emit_left and not emit_right else "ONLY IN UAT")
        )

        # 1) PROD row
        if emit_left:
            ws.append([cfg.left_label] + [left_vals.get(c, "") for c in display_cols] +
                      [rec_status, ("YES" if (both_present and all_equal) else "")])
            for j, cell in enumerate(ws[ws.max_row], start=1):
                cell.fill = fill_left
                if 1 < j <= len(display_cols) + 1:
                    colname = header[j - 1]
                    write_cell(cell, left_vals.get(colname, ""), colname)

        # 2) UAT row
        if emit_right:
            ws.append([cfg.right_label] + [right_vals.get(c, "") for c in display_cols] +
                      [rec_status, ("YES" if (both_present and all_equal) else "NO" if both_present else "")])
            row_ref = ws.max_row
            for j in range(2, len(display_cols) + 2):
                colname = header[j - 1]
                c2 = ws[row_ref][j - 1]
                write_cell(c2, right_vals.get(colname, ""), colname)

                if emit_left and both_present:
                    a = r.get(f"{colname}_f1", "")
                    b = r.get(f"{colname}_f2", "")
                    c1 = ws[row_ref - 1][j - 1]
                    c1.fill = fill_left
                    c2.fill = fill_right
                    if not cell_equal(a, b, colname, numeric_cols, date_cols, cfg):
                        c1.fill = fill_mismatch
                        c2.fill = fill_mismatch

            # Validation color for UAT row (the row that carries YES/NO when both present)
            vcell = ws[row_ref][len(display_cols) + 2]
            if both_present:
                vcell.fill = (fill_yes if all_equal else fill_no)

        # If only PROD row emitted (no UAT), color its Validation cell (blank but colored?)
        if emit_left and not emit_right:
            row_ref = ws.max_row
            vcell = ws[row_ref][len(display_cols) + 2]
            # Leave it blank; optionally color to draw attention (commented)
            # vcell.fill = fill_no
            pass

    # Freeze header, filter, auto-widths
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in list(col_cells)[:5000])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

    # Optional: Summary sheet
    try:
        ws2 = wb.create_sheet("Summary")
        total_rows = merged.shape[0]
        only_prod = 0
        only_uat = 0
        matches = 0
        mismatches = 0
        for _, r in merged.iterrows():
            left_missing_row = all(is_missing(r.get(f"{c}_f1", ""), cfg) for c in display_cols)
            right_missing_row = all(is_missing(r.get(f"{c}_f2", ""), cfg) for c in display_cols)
            if left_missing_row and (not right_missing_row):
                only_uat += 1
                continue
            if right_missing_row and (not left_missing_row):
                only_prod += 1
                continue
            # both present
            all_equal = True
            for col in display_cols:
                if not cell_equal(r.get(f"{col}_f1", ""), r.get(f"{col}_f2", ""), col, numeric_cols, date_cols, cfg):
                    all_equal = False; break
            if all_equal:
                matches += 1
            else:
                mismatches += 1

        ws2.append(["Metric", "Value"])
        for k, v in [
            ("Keys used", ", ".join(keys_used) if keys_used else "Row order"),
            ("Total pairs", total_rows),
            ("Matches (both present)", matches),
            ("Mismatches (both present)", mismatches),
            ("Only in PROD", only_prod),
            ("Only in UAT", only_uat),
        ]:
            ws2.append([k, v])
        for cell in ws2[1]:
            cell.font = Font(bold=True); cell.fill = header_fill
        ws2.auto_filter.ref = f"A1:B{ws2.max_row}"
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 22
    except Exception:
        pass

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def compare_and_render(
    left_source: Any,      # PROD
    right_source: Any,     # UAT
    cfg: CompareConfig,
    explicit_keys: Optional[List[str]] = None,
    sheet_left: Optional[str] = None,
    sheet_right: Optional[str] = None,
) -> Dict[str, Any]:
    df1 = load_any(left_source, cfg, sheet_left)
    df2 = load_any(right_source, cfg, sheet_right)
    if not isinstance(df1, pd.DataFrame) or not isinstance(df2, pd.DataFrame):
        raise ValueError("Failed to load input files as DataFrames.")

    all_cols = sorted(set(df1.columns) | set(df2.columns))

    # Keys
    keys = None
    if explicit_keys:
        keys = [k.strip().lower() if cfg.case_insensitive_headers else k.strip() for k in explicit_keys]
        for k in keys:
            if k not in df1.columns or k not in df2.columns:
                raise ValueError(f"Key column '{k}' not present in both files.")
    else:
        keys = pick_keys_auto(df1, df2, cfg)

    df1k, keys_used = add_key_and_seq(df1, keys)
    df2k, _ = add_key_and_seq(df2, keys)

    display_cols = (keys_used or []) + [c for c in all_cols if c not in (keys_used or [])]
    numeric_cols = detect_numeric_cols(df1, df2, exclude=(keys_used or []), cfg=cfg)
    date_cols = detect_date_cols(df1, df2, exclude=(keys_used or []), cfg=cfg)

    merged = build_pairs(df1k, df2k)

    # Compact preview
    preview_rows = []
    for _, r in merged.head(cfg.limit_preview_rows).iterrows():
        left_missing_row = all(is_missing(r.get(f"{c}_f1", ""), cfg) for c in display_cols)
        right_missing_row = all(is_missing(r.get(f"{c}_f2", ""), cfg) for c in display_cols)
        both_present = (not left_missing_row) and (not right_missing_row)
        all_equal = True
        diffs = []
        if both_present:
            for col in display_cols:
                a = r.get(f"{col}_f1", "")
                b = r.get(f"{col}_f2", "")
                if not cell_equal(a, b, col, numeric_cols, date_cols, cfg):
                    all_equal = False
                    diffs.append(col)
        status = "BOTH" if both_present else ("ONLY IN PROD" if not left_missing_row else "ONLY IN UAT")
        preview_rows.append({
            "__key__": r["__key__"], "__seq__": r["__seq__"],
            "RecordStatus": status,
            "Match": ("YES" if (both_present and all_equal) else ("NO" if both_present else "")),
            "Diff_Cols": ", ".join(diffs)
        })
    preview_df = pd.DataFrame(preview_rows)

    excel_bytes = write_excel(
        merged=merged,
        display_cols=display_cols,
        numeric_cols=numeric_cols,
        date_cols=date_cols,
        keys_used=keys_used,
        cfg=cfg,
    )

    return {
        "summary": {
            "keys_used": keys_used,
            "display_cols": display_cols,
            "numeric_cols": numeric_cols,
            "date_cols": date_cols,
            "preview_count": len(preview_df),
        },
        "preview_df": preview_df,
        "excel_bytes": excel_bytes,
    }
